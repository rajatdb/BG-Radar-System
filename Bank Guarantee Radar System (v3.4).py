import sys
import io
import re
import urllib.request
import threading
import time
import json
import os
import getpass
import warnings
from datetime import datetime, timedelta
import pandas as pd

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from PIL import Image, ImageDraw, ImageTk, ImageFont
from auto_version_updater import InAppSeamlessUpdater
from user_security_tracker import DeviceUserSecurityTracker

# openpyxl for Excel styling
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Suppress Pandas Date Parsing UserWarnings
warnings.filterwarnings("ignore", category=UserWarning, module="pandas")

# --- Current Application Version & Credentials ---
CURRENT_VERSION = "v3.4"

SHEET_ID = "1dx8Hs5lp2LAIj914XFbA3NAvZQ_j1pmJbzyQsJ_kTOE"
GID_MAIN = "161119154"

CSV_MAIN_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=csv&gid={GID_MAIN}"
CSV_HISTORY_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/gviz/tq?tqx=out:csv&sheet=UPDATE_HISTORY"
# --- Google Apps Script Web App URL for User Tracking ---
WEB_APP_URL = "https://script.google.com/macros/s/AKfycbwpECn1HQa-Q3Jo6ivC87UTr8RRUuvUjA15JP58gCC55lut7ZvfAzRzf6BQ2ovVvjmFsQ/exec"

# Base Directory Resolution
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

DELETED_LOGS_FILE = os.path.join(BASE_DIR, "bg_deleted_history.json")
KNOWN_ENTRIES_FILE = os.path.join(BASE_DIR, "bg_known_entries.json")


def get_active_system_username():
    """Fetches currently logged-in Windows OS username for security logging"""
    try:
        user = getpass.getuser()
        return user.upper() if user else "OFFICIAL_STAFF"
    except Exception:
        return os.environ.get("USERNAME", "OFFICIAL_STAFF").upper()


def load_json_set(file_path):
    """Utility to load persistent sets from JSON"""
    if os.path.exists(file_path):
        try:
            with open(file_path, "r") as f:
                return set(json.load(f))
        except Exception:
            return set()
    return set()


def save_json_set(file_path, data_set):
    """Utility to save persistent sets into JSON"""
    try:
        with open(file_path, "w") as f:
            json.dump(list(data_set), f, indent=2)
    except Exception:
        pass


def fix_utf8_encoding(raw_text):
    """Fixes garbled unicode characters like 'â‚¹' and restores '₹'"""
    if not raw_text:
        return ""
    
    replacements = {
        'â‚¹': '₹',
        'Â': '',
        'â‚': '₹',
        'â': ''
    }
    for bad_char, good_char in replacements.items():
        raw_text = raw_text.replace(bad_char, good_char)
        
    return raw_text


def clean_cell_value(val):
    """Cleans empty string, 'nan', and None values into structured display text"""
    if pd.isna(val) or val is None:
        return "N/A"
    s = str(val).strip()
    if s == "" or s.lower() == "nan":
        return "N/A"
    return fix_utf8_encoding(s)


def format_indian_date(val_str):
    """Converts raw date strings OR Excel Serial Numbers to clean DD/MM/YYYY"""
    cleaned = clean_cell_value(val_str)
    if cleaned == "N/A" or cleaned == "":
        return "N/A"
    
    cleaned = cleaned.lstrip("'").strip()

    if cleaned.replace('.', '', 1).isdigit():
        try:
            num = float(cleaned)
            if 30000 < num < 70000:
                excel_base = datetime(1899, 12, 30)
                dt_obj = excel_base + timedelta(days=num)
                return dt_obj.strftime('%d/%m/%Y')
        except Exception:
            pass

    try:
        cleaned_format = cleaned.replace('-', '/')
        if re.match(r'^\d{4}/\d{2}/\d{2}', cleaned_format):
            dt = pd.to_datetime(cleaned_format, dayfirst=False, errors='coerce')
        else:
            dt = pd.to_datetime(cleaned_format, dayfirst=True, errors='coerce')

        if pd.isna(dt):
            return cleaned
        return dt.strftime('%d/%m/%Y')
    except Exception:
        return cleaned


def format_indian_datetime(val_str):
    """Formats timestamp string into readable IST Format (DD/MM/YYYY hh:mm AM/PM)"""
    cleaned = clean_cell_value(val_str)
    if cleaned == "N/A" or cleaned == "":
        return "N/A"
    try:
        cleaned_format = cleaned.replace('-', '/')
        if re.match(r'^\d{4}/\d{2}/\d{2}', cleaned_format):
            dt = pd.to_datetime(cleaned_format, dayfirst=False, errors='coerce')
        else:
            dt = pd.to_datetime(cleaned_format, dayfirst=True, errors='coerce')

        if pd.isna(dt):
            return cleaned
        return dt.strftime('%d/%m/%Y %I:%M %p')
    except Exception:
        return cleaned


def format_indian_currency(val):
    """Formats raw numbers into uniform Indian Rupee Currency (₹ XX,XX,XXX)"""
    cleaned_val = clean_cell_value(val)
    if cleaned_val == "N/A":
        return "N/A"

    cleaned_num = re.sub(r'[^\d.]', '', cleaned_val)

    try:
        amount = float(cleaned_num)
        amount_int = int(round(amount))
        s = str(amount_int)
        if len(s) <= 3:
            formatted = s
        else:
            last_three = s[-3:]
            remaining = s[:-3]
            groups = []
            while len(remaining) > 2:
                groups.insert(0, remaining[-2:])
                remaining = remaining[:-2]
            if remaining:
                groups.insert(0, remaining)
            formatted = ",".join(groups) + "," + last_three

        return f"₹ {formatted}"
    except ValueError:
        return cleaned_val


def fetch_sheet_stream_fast():
    """Fetches Main Radar Data and creates Master Dictionary for cross-tab auto-fill"""
    req = urllib.request.Request(
        CSV_MAIN_URL, 
        headers={'User-Agent': 'Mozilla/5.0'}
    )
    with urllib.request.urlopen(req, timeout=8) as response:
        raw_bytes = response.read()
        try:
            csv_data = raw_bytes.decode('utf-8-sig')
        except UnicodeDecodeError:
            csv_data = raw_bytes.decode('latin-1')

    if not csv_data:
        raise Exception("Unable to retrieve live sheet stream.")

    csv_data = fix_utf8_encoding(csv_data)
    df = pd.read_csv(io.StringIO(csv_data))
    
    if df.empty:
        raise Exception("Google Sheet returned no active records.")

    contractor_col = df.columns[1] if len(df.columns) > 1 else df.columns[0]
    bg_no_col      = df.columns[4] if len(df.columns) > 4 else df.columns[0]
    amt_col        = df.columns[5] if len(df.columns) > 5 else df.columns[0]
    date_col       = df.columns[6] if len(df.columns) > 6 else df.columns[0]
    aden_col       = df.columns[10] if len(df.columns) > 10 else df.columns[0]
    dealer_col     = df.columns[11] if len(df.columns) > 11 else df.columns[0]
    remarks_col    = df.columns[13] if len(df.columns) > 13 else None

    for col in df.columns:
        if df[col].dtype == 'object':
            df[col] = df[col].astype(str).apply(clean_cell_value)

    df['Parsed_Expiry'] = pd.to_datetime(df[date_col].astype(str).str.replace('-', '/'), errors='coerce', dayfirst=True)
    today = datetime.now()
    df['Days_Remaining'] = (df['Parsed_Expiry'] - today).dt.days

    df['Formatted_Amount'] = df[amt_col].apply(format_indian_currency) if amt_col else "N/A"
    df['Clean_Remarks'] = df[remarks_col].apply(lambda x: clean_cell_value(x).upper()) if remarks_col else "N/A"

    df['Radar_Warning'] = 'VALID / OK'
    df.loc[df['Days_Remaining'] <= 90, 'Radar_Warning'] = 'DUE FOR RENEWAL (< 90 DAYS)'
    df.loc[df['Days_Remaining'] <= 60, 'Radar_Warning'] = 'HIGH PRIORITY ACTION (< 60 DAYS)'
    df.loc[df['Days_Remaining'] <= 30, 'Radar_Warning'] = 'URGENT ACTION REQUIRED (< 30 DAYS)'
    df.loc[df['Days_Remaining'] < 0, 'Radar_Warning'] = 'CRITICAL OVERDUE (NO REMARKS)'

    has_remarks_mask = (df['Clean_Remarks'] != "N/A") & (df['Clean_Remarks'] != "")
    df.loc[has_remarks_mask, 'Radar_Warning'] = '✅ RESOLVED / REMARKS ADDED'

    df_active_radar = df[df['Radar_Warning'] != '✅ RESOLVED / REMARKS ADDED'].copy()
    df_resolved_bg = df[df['Radar_Warning'] == '✅ RESOLVED / REMARKS ADDED'].copy()

    master_bg_lookup = {}
    for _, r in df.iterrows():
        bg_key = clean_cell_value(r[bg_no_col])
        if bg_key != "N/A":
            master_bg_lookup[bg_key] = {
                'aden': clean_cell_value(r[aden_col]),
                'dealer': clean_cell_value(r[dealer_col]),
                'contractor': clean_cell_value(r[contractor_col])
            }

    return df_active_radar, df_resolved_bg, master_bg_lookup, date_col, aden_col, dealer_col, contractor_col, bg_no_col, 'Formatted_Amount', 'Clean_Remarks'


def fetch_history_stream(master_bg_lookup):
    """Fetches Edit History and cross-fills missing ADEN/Dealer from Master Lookup"""
    try:
        req = urllib.request.Request(CSV_HISTORY_URL, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=5) as response:
            csv_data = response.read().decode('utf-8-sig')
            df_hist = pd.read_csv(io.StringIO(csv_data))
            
            if df_hist.empty:
                return pd.DataFrame()

            deleted_set = load_json_set(DELETED_LOGS_FILE)
            valid_rows = []

            for _, row in df_hist.iterrows():
                vals = [clean_cell_value(v) for v in row]
                if len(vals) >= 9:
                    bg_no = vals[4] if len(vals) > 4 else "N/A"
                    new_extended_date = vals[7] if len(vals) > 7 else "N/A"
                    updated_on_str = vals[8]
                    
                    signature = f"{bg_no}_{new_extended_date}_{updated_on_str}"
                    if signature in deleted_set:
                        continue

                    if bg_no in master_bg_lookup:
                        m_info = master_bg_lookup[bg_no]
                        if vals[1] == "N/A" and m_info['aden'] != "N/A":
                            vals[1] = m_info['aden']
                        if vals[2] == "N/A" and m_info['dealer'] != "N/A":
                            vals[2] = m_info['dealer']
                        if vals[3] == "N/A" and m_info['contractor'] != "N/A":
                            vals[3] = m_info['contractor']

                    valid_rows.append(vals)

            return pd.DataFrame(valid_rows) if valid_rows else pd.DataFrame()
    except Exception:
        return pd.DataFrame()


class ExecutiveRailwaysPortal(tk.Tk):
    def __init__(self):
        super().__init__()
        self.active_user = get_active_system_username()
        self.title(f"INDIAN RAILWAYS — Bank Guarantee Radar System ({CURRENT_VERSION}) | User: {self.active_user}")
        self.geometry("1500x900")
        self.minsize(1150, 720)

        try:
            self.state('zoomed')
        except Exception:
            pass

        self.NAVY = "#061d33"
        self.AMBER = "#d97706"
        self.BG_CANVAS = "#f1f5f9"
        self.TEXT_DARK = "#0f172a"
        self.TEXT_MUTED = "#64748b"
        self.BORDER_LINE = "#334155"
        self.BORDER_LIGHT = "#94a3b8"

        self.configure(bg=self.BG_CANVAS)
        self.raw_warnings_df = pd.DataFrame()
        self.df_resolved_cached = pd.DataFrame()
        self.df_history_cached = pd.DataFrame()
        
        self.current_filtered_active_df = pd.DataFrame()
        self.current_filtered_resolved_df = pd.DataFrame()
        self.current_filtered_history_df = pd.DataFrame()

        self.is_syncing = False
        self.selected_history_item = None
        self.deleted_blacklist = load_json_set(DELETED_LOGS_FILE)
        self.known_entries = load_json_set(KNOWN_ENTRIES_FILE)
        self.new_logs_detected = set()

        self.is_blinking = False
        self.blink_state = False

        # Global Deselect Click Event Listener
        self.bind("<Button-1>", self.global_click_deselect)

        # 1. Header Frame
        header = tk.Frame(self, bg=self.NAVY, height=90, highlightbackground=self.BORDER_LINE, highlightthickness=1)
        header.pack(fill="x", padx=10, pady=(10, 5))

        LOGO_SIZE = 68
        self.logo_canvas = tk.Canvas(header, width=LOGO_SIZE, height=LOGO_SIZE, bg=self.NAVY, highlightthickness=0, bd=0)
        self.logo_canvas.pack(side="left", padx=(20, 10), pady=10)

        self.logo_img = self.load_custom_badge_image(size=(LOGO_SIZE, LOGO_SIZE))
        if self.logo_img:
            self.logo_canvas.create_image(LOGO_SIZE // 2, LOGO_SIZE // 2, image=self.logo_img)
        else:
            scale = 4
            sz = LOGO_SIZE * scale
            img = Image.new("RGBA", (sz, sz), (0, 0, 0, 0))
            draw = ImageDraw.Draw(img)
            draw.ellipse([int(4*scale), int(4*scale), int((LOGO_SIZE-4)*scale), int((LOGO_SIZE-4)*scale)], outline=self.AMBER, width=int(2.5*scale))
            draw.ellipse([int(9*scale), int(9*scale), int((LOGO_SIZE-9)*scale), int((LOGO_SIZE-9)*scale)], outline="#FFFFFF", width=int(1*scale))
            img_smooth = img.resize((LOGO_SIZE, LOGO_SIZE), Image.Resampling.LANCZOS)
            self.logo_img_fallback = ImageTk.PhotoImage(img_smooth)
            self.logo_canvas.create_image(LOGO_SIZE // 2, LOGO_SIZE // 2, image=self.logo_img_fallback)
            self.logo_canvas.create_text(LOGO_SIZE // 2, LOGO_SIZE // 2, text="🚆", font=("Segoe UI Emoji", 22), fill="#ffffff")

        title_frame = tk.Frame(header, bg=self.NAVY)
        title_frame.pack(side="left", padx=10, pady=12)

        logo_label = tk.Label(
            title_frame, text="INDIAN RAILWAYS - BIKANER DIVISION",
            font=("Segoe UI", 16, "bold"), fg=self.AMBER, bg=self.NAVY
        )
        logo_label.pack(anchor="w")

        sub_header = tk.Label(
            title_frame, text="BANK GUARANTEE RADAR & EXPIRY MONITORING PORTAL",
            font=("Segoe UI", 11, "bold"), fg="#ffffff", bg=self.NAVY
        )
        sub_header.pack(anchor="w")

        # Security User Tag in Header
        user_badge_frame = tk.Frame(header, bg="#0f2942", highlightbackground=self.AMBER, highlightthickness=1)
        user_badge_frame.pack(side="right", padx=20, pady=18)

        lbl_user_icon = tk.Label(user_badge_frame, text=f"👤 Staff User: {self.active_user}", font=("Segoe UI", 10, "bold"), fg="#38bdf8", bg="#0f2942", padx=10, pady=5)
        lbl_user_icon.pack()

        # 2. Executive Stat Summary Cards Container
        self.cards_container = tk.Frame(self, bg=self.BG_CANVAS, highlightbackground=self.BORDER_LIGHT, highlightthickness=1)
        self.cards_container.pack(fill="x", padx=10, pady=5)

        self.cards_frame = tk.Frame(self.cards_container, bg=self.BG_CANVAS)
        self.cards_frame.pack(fill="x", padx=5, pady=5)

        self.card_critical = self.create_card(self.cards_frame, "🛑 CRITICAL OVERDUE (NO REMARKS)", "0", "#991b1b", "#ffe4e6", 0, filter_key="CRITICAL OVERDUE")
        self.card_red = self.create_card(self.cards_frame, "🚨 URGENT ACTION (< 30 DAYS)", "0", "#dc2626", "#fef2f2", 1, filter_key="URGENT")
        self.card_orange = self.create_card(self.cards_frame, "⚠️ HIGH PRIORITY (< 60 DAYS)", "0", "#ea580c", "#fff7ed", 2, filter_key="HIGH PRIORITY")
        self.card_yellow = self.create_card(self.cards_frame, "⏳ DUE RENEWAL (< 90 DAYS)", "0", "#ca8a04", "#fef9c3", 3, filter_key="DUE FOR RENEWAL")
        self.card_total = self.create_card(self.cards_frame, "🛡️ TOTAL PENDING RADARS", "0", "#0284c7", "#f0f9ff", 4, filter_key="ALL")

        # 3. Action Control Bar
        control_bar = tk.Frame(self, bg=self.NAVY, height=55, highlightbackground=self.BORDER_LINE, highlightthickness=1)
        control_bar.pack(fill="x", padx=10, pady=5)

        self.refresh_btn = tk.Button(
            control_bar, text="🔄 Live Sync", font=("Segoe UI", 10, "bold"),
            bg="#059669", fg="#ffffff",
            activebackground="#047857", activeforeground="#ffffff",
            disabledforeground="#ffffff",
            relief="flat", bd=0, highlightthickness=0,
            cursor="hand2", command=self.trigger_async_load
        )
        self.refresh_btn.pack(side="left", padx=12, pady=8, ipady=5, ipadx=14)

        export_btn = tk.Button(
            control_bar, text="📁 Export to Excel (.xlsx)", font=("Segoe UI", 10, "bold"),
            bg="#2563eb", fg="#ffffff",
            activebackground="#1d4ed8", activeforeground="#ffffff",
            disabledforeground="#ffffff",
            relief="flat", bd=0, highlightthickness=0,
            cursor="hand2", command=self.export_excel
        )
        export_btn.pack(side="left", padx=5, pady=8, ipady=5, ipadx=12)
	
	# Connected Users Button
        users_btn = tk.Button(
            control_bar, text="👥 Live Active Staff", font=("Segoe UI", 10, "bold"),
            bg="#0284c7", fg="#ffffff",
            activebackground="#0369a1", activeforeground="#ffffff",
            relief="flat", bd=0, highlightthickness=0,
            cursor="hand2", command=self.show_connected_users_dialog
        )
        users_btn.pack(side="left", padx=5, pady=8, ipady=5, ipadx=12)

        search_lbl = tk.Label(control_bar, text="🔍 Universal Search:", font=("Segoe UI", 10, "bold"), fg="#ffffff", bg=self.NAVY)
        search_lbl.pack(side="left", padx=(15, 5))

        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", self.filter_table)

        search_entry = tk.Entry(
            control_bar, textvariable=self.search_var, font=("Segoe UI", 10),
            bg="#ffffff", fg="#000000", insertbackground="#000000", width=25,
            relief="solid", bd=1, highlightthickness=0
        )
        search_entry.pack(side="left", padx=5, ipady=4)
        search_entry.bind("<KeyRelease>", self.filter_table)

        self.status_label = tk.Label(
            control_bar, text="Connecting Fast Stream...",
            font=("Segoe UI", 10, "italic"), fg="#cbd5e1", bg=self.NAVY
        )
        self.status_label.pack(side="right", padx=15)

        # Footer Bar
        footer = tk.Frame(self, bg=self.NAVY, height=35, highlightbackground=self.BORDER_LINE, highlightthickness=1)
        footer.pack(fill="x", side="bottom", padx=10, pady=(5, 10))

        footer_left = tk.Label(
            footer, text=f"OFFICIAL USE ONLY — NORTH WESTERN RAILWAY | LOGGED AS: {self.active_user}",
            font=("Segoe UI", 9, "bold"), fg="#94a3b8", bg=self.NAVY
        )
        footer_left.pack(side="left", padx=20, pady=6)

        footer_right = tk.Label(
            footer, text="Designed & Developed by Shri Rajat Dubey (Sr. Accountant)",
            font=("Segoe UI", 9, "bold"), fg=self.AMBER, bg=self.NAVY
        )
        footer_right.pack(side="right", padx=20, pady=6)

        # 4. TABBED CONTAINER
        content_frame = tk.Frame(self, bg=self.BG_CANVAS)
        content_frame.pack(fill="both", expand=True, padx=10, pady=5)

        self.inspector = tk.Frame(
            content_frame, bg="#ffffff", width=330,
            highlightbackground=self.BORDER_LINE, highlightthickness=1
        )
        self.inspector.pack(side="right", fill="y", padx=(10, 0))
        self.inspector.pack_propagate(False)

        insp_header = tk.Frame(self.inspector, bg=self.NAVY, highlightbackground=self.BORDER_LINE, highlightthickness=1)
        insp_header.pack(fill="x")

        insp_title = tk.Label(insp_header, text="📋 Record Inspector", font=("Segoe UI", 11, "bold"), fg=self.AMBER, bg=self.NAVY, pady=8)
        insp_title.pack(anchor="w", padx=12)

        self.insp_text = tk.Text(
            self.inspector, font=("Segoe UI", 10), fg=self.TEXT_DARK, bg="#ffffff",
            wrap="word", bd=0, highlightthickness=0, padx=12, pady=12
        )
        self.insp_text.insert("1.0", "Click on any record row to inspect full details.")
        self.insp_text.config(state="disabled")
        self.insp_text.pack(fill="both", expand=True)

        self.btn_delete_frame = tk.Frame(self.inspector, bg="#ffffff", height=50)
        self.btn_delete_frame.pack(fill="x", side="bottom", padx=12, pady=15)
        self.btn_delete_frame.pack_propagate(False)

        self.delete_btn = tk.Button(
            self.btn_delete_frame, text="🗑️ Delete Selected History Entry", font=("Segoe UI", 10, "bold"),
            bg="#dc2626", fg="#ffffff", activebackground="#b91c1c", activeforeground="#ffffff",
            relief="flat", bd=0, highlightthickness=0, cursor="hand2", command=self.delete_history_entry
        )
        self.delete_btn.pack_forget()

        self.notebook = ttk.Notebook(content_frame)
        self.notebook.pack(side="left", fill="both", expand=True)

        self.tab_active = tk.Frame(self.notebook, bg=self.NAVY)
        self.notebook.add(self.tab_active, text="📋 Pending Active BG Radar ")

        self.tab_resolved = tk.Frame(self.notebook, bg=self.NAVY)
        self.notebook.add(self.tab_resolved, text=" ✅ Validated BGs (Remarks) ")

        self.tab_history = tk.Frame(self.notebook, bg=self.NAVY)
        self.notebook.add(self.tab_history, text=" 📜 Update History ")

        self.table_frame = tk.Frame(self.tab_active, bg=self.NAVY, highlightbackground=self.BORDER_LINE, highlightthickness=1)
        self.table_frame.pack(fill="both", expand=True)

        style = ttk.Style()
        style.theme_use("clam")

        style.configure(
            "Treeview",
            background="#ffffff",
            foreground="#0f172a",
            rowheight=48,
            fieldbackground="#ffffff",
            font=("Segoe UI", 10),
            bordercolor=self.BORDER_LINE,
            borderwidth=1
        )

        style.configure(
            "Treeview.Heading",
            background=self.NAVY,
            foreground="#ffffff",
            font=("Segoe UI", 10, "bold"),
            bordercolor=self.BORDER_LINE,
            relief="solid",
            borderwidth=1
        )

        style.map("Treeview", 
                  background=[("selected", "#1e3a8a"), ("focus", "#1e3a8a")], 
                  foreground=[("selected", "#ffffff"), ("focus", "#ffffff")])

        style.map("Treeview.Heading",
                  background=[("active", self.NAVY), ("pressed", self.NAVY)],
                  foreground=[("active", self.AMBER), ("pressed", self.AMBER)])

        self.tree = ttk.Treeview(
            self.table_frame,
            columns=("ADEN_Dept", "Dealer", "Contractor", "BG_No_Date", "Amount", "Expiry_Date", "Days_Left", "Radar"),
            show="headings"
        )

        self.tree.heading("ADEN_Dept", text="ADEN / Dept.")
        self.tree.heading("Dealer", text="Dealer Name")
        self.tree.heading("Contractor", text="Contractor / Firm Name")
        self.tree.heading("BG_No_Date", text="Bank Guarantee No.")
        self.tree.heading("Amount", text="BG Amount")
        self.tree.heading("Expiry_Date", text="Validity Date")
        self.tree.heading("Days_Left", text="Days Left")
        self.tree.heading("Radar", text="Executive Radar Status")

        self.tree.column("ADEN_Dept", width=120, minwidth=100, anchor="center", stretch=True)
        self.tree.column("Dealer", width=130, minwidth=110, anchor="center", stretch=True)
        self.tree.column("Contractor", width=220, minwidth=160, anchor="w", stretch=True)
        self.tree.column("BG_No_Date", width=160, minwidth=130, anchor="w", stretch=True)
        self.tree.column("Amount", width=125, minwidth=100, anchor="e", stretch=True)
        self.tree.column("Expiry_Date", width=110, minwidth=90, anchor="center", stretch=True)
        self.tree.column("Days_Left", width=85, minwidth=70, anchor="center", stretch=True)
        self.tree.column("Radar", width=220, minwidth=160, anchor="center", stretch=True)

        scrollbar = ttk.Scrollbar(self.table_frame, orient="vertical", command=self.tree.yview)
        h_scrollbar = ttk.Scrollbar(self.table_frame, orient="horizontal", command=self.tree.xview)
        
        self.tree.configure(yscrollcommand=scrollbar.set, xscrollcommand=h_scrollbar.set)

        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        h_scrollbar.pack(side="bottom", fill="x")

        self.table_frame.bind("<Configure>", self.autofit_and_wrap)

        self.tree.tag_configure("CRITICAL_OVERDUE", background="#ffe4e6", foreground="#991b1b")
        self.tree.tag_configure("URGENT", background="#fee2e2", foreground="#991b1b")
        self.tree.tag_configure("HIGH_PRIORITY", background="#ffedd5", foreground="#9a3412")
        self.tree.tag_configure("RENEWAL_DUE", background="#fef9c3", foreground="#854d0e")

        self.tree.bind("<<TreeviewSelect>>", self.on_row_select)
        self.tree.bind("<Button-3>", lambda e: self.show_copy_menu(e, self.tree))

        self.setup_resolved_table(self.tab_resolved)
        self.setup_history_table(self.tab_history)

        self.notebook.bind("<<NotebookTabChanged>>", self.on_tab_changed)

        self.after(50, self.trigger_async_load)
        # Check for Software Version Updates
        self.after(1500, self.trigger_live_seamless_update)
	# Initialize Security Tracker & Start Background Ping
        self.security_tracker = DeviceUserSecurityTracker()
        self.start_heartbeat_ping()

    # ------------------------------------------------------------------
    # Ensure these methods are defined INSIDE ExecutiveRailwaysPortal class
    # ------------------------------------------------------------------
    def send_heartbeat_ping(self):
        """Sends periodic live status ping to Google Sheet Web App"""
        if not WEB_APP_URL or "YOUR_DEPLOYMENT_ID" in WEB_APP_URL:
            return
        try:
            payload = self.security_tracker.get_device_info()
            data_bytes = json.dumps(payload).encode('utf-8')
            req = urllib.request.Request(
                WEB_APP_URL, 
                data=data_bytes, 
                headers={'Content-Type': 'application/json'}
            )
            urllib.request.urlopen(req, timeout=5)
        except Exception:
            pass

    def start_heartbeat_ping(self):
        """Triggers ping every 45 seconds on background thread"""
        def ping_loop():
            while True:
                self.send_heartbeat_ping()
                time.sleep(45)
        
        threading.Thread(target=ping_loop, daemon=True).start()

    def show_connected_users_dialog(self):
        """Displays dialog showing all registered and live online staff members"""
        if not WEB_APP_URL or "YOUR_DEPLOYMENT_ID" in WEB_APP_URL:
            messagebox.showwarning("Configuration Needed", "Please set your deployed Google Web App URL in WEB_APP_URL variable.")
            return

        dialog = tk.Toplevel(self)
        dialog.title("🌐 Connected Network Staff Users")
        dialog.geometry("700x420")
        dialog.configure(bg=self.NAVY)
        dialog.transient(self)
        dialog.grab_set()

        lbl_title = tk.Label(
            dialog, text="👥 Network Staff Activity Dashboard", 
            font=("Segoe UI", 12, "bold"), fg=self.AMBER, bg=self.NAVY
        )
        lbl_title.pack(pady=10)

        frame = tk.Frame(dialog, bg=self.NAVY)
        frame.pack(fill="both", expand=True, padx=15, pady=10)

        tree = ttk.Treeview(
            frame, 
            columns=("Computer", "User", "IP", "Status", "Last_Active"), 
            show="headings"
        )
        
        tree.heading("Computer", text="Computer Name")
        tree.heading("User", text="Staff User")
        tree.heading("IP", text="IP Address")
        tree.heading("Status", text="Live Status")
        tree.heading("Last_Active", text="Last Seen Timestamp")

        tree.column("Computer", width=140, anchor="center")
        tree.column("User", width=120, anchor="center")
        tree.column("IP", width=110, anchor="center")
        tree.column("Status", width=110, anchor="center")
        tree.column("Last_Active", width=160, anchor="center")

        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)

        tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        tree.tag_configure("ONLINE", background="#dcfce7", foreground="#166534")
        tree.tag_configure("OFFLINE", background="#fee2e2", foreground="#991b1b")

        def format_to_ist_display(raw_time_str):
            """Converts raw ISO UTC timestamp strings to clean DD/MM/YYYY hh:mm AM/PM IST"""
            if not raw_time_str or raw_time_str == "N/A":
                return "N/A"
            try:
                # Parse ISO date string
                dt = pd.to_datetime(raw_time_str, errors='coerce')
                if pd.isna(dt):
                    return str(raw_time_str)
                
                # Convert UTC / Naive datetime to Indian Standard Time (+5:30)
                if dt.tzinfo is None or dt.tzinfo.utcoffset(dt) is None:
                    dt = dt.tz_localize('UTC').tz_convert('Asia/Kolkata')
                else:
                    dt = dt.tz_convert('Asia/Kolkata')
                    
                return dt.strftime('%d/%m/%Y %I:%M %p')
            except Exception:
                return str(raw_time_str)

        def fetch_users():
            try:
                req = urllib.request.Request(WEB_APP_URL, headers={'User-Agent': 'Mozilla/5.0'})
                with urllib.request.urlopen(req, timeout=8) as resp:
                    users = json.loads(resp.read().decode('utf-8'))
                    
                    dialog.after(0, lambda: [tree.delete(item) for item in tree.get_children()])
                    
                    for u in users:
                        status_str = "🟢 ONLINE" if u.get('status') == "ONLINE" else "🔴 OFFLINE"
                        tag = u.get('status', 'OFFLINE')
                        
                        # Formatted readable IST timestamp
                        formatted_last_seen = format_to_ist_display(u.get('last_seen', 'N/A'))
                        
                        dialog.after(0, lambda u=u, s=status_str, t=tag, ls=formatted_last_seen: tree.insert("", "end", values=(
                            u.get('computer_name', 'N/A'),
                            u.get('username', 'N/A'),
                            u.get('ip_address', 'N/A'),
                            s,
                            ls
                        ), tags=(t,)))
            except Exception as e:
                dialog.after(0, lambda: messagebox.showerror("Error", f"Failed to fetch active users:\n{e}", parent=dialog))

        threading.Thread(target=fetch_users, daemon=True).start()

    def trigger_live_seamless_update(self):
        """Triggers background seamless version checker"""
        updater = InAppSeamlessUpdater(
            current_version=CURRENT_VERSION,
            product_code="NWR_BG_RADAR_PRO",
            parent_app=self
        )
        updater.check_and_update()

    def global_click_deselect(self, event):
        """Deselects row only if click is outside Treeview, Inspector Panel, and Delete Button"""
        w = event.widget
        if w in (self.delete_btn, self.insp_text, self.inspector, self.btn_delete_frame):
            return
        if not isinstance(w, ttk.Treeview):
            self.clear_all_table_selections()

    def clear_all_table_selections(self):
        """Clears selection state across all 3 tables & resets Inspector Panel while keeping tag colors intact"""
        for t in [self.tree, self.resolved_tree, self.history_tree]:
            selected = t.selection()
            if selected:
                t.selection_remove(selected)
        self.selected_history_item = None
        self.delete_btn.pack_forget()
        self.set_inspector_text("Click on any record row to inspect full details.")

    def load_custom_badge_image(self, size=(68, 68)):
        possible_names = ["app_icon.ico", "logo.png", "logo.jpg", "logo.jpeg", "badge.png", "badge.jpg", "Picsart_26-08-04_16-02-35-718.jpg"]
        for fname in possible_names:
            fpath = os.path.join(BASE_DIR, fname)
            if os.path.exists(fpath):
                try:
                    pil_img = Image.open(fpath).convert("RGBA")
                    pil_img = pil_img.resize(size, Image.Resampling.LANCZOS)
                    return ImageTk.PhotoImage(pil_img)
                except Exception:
                    pass
        return None

    def start_tab_blinking(self):
        if not self.is_blinking:
            self.is_blinking = True
            self._blink_loop()

    def stop_tab_blinking(self):
        self.is_blinking = False
        self.notebook.tab(self.tab_history, text=" 📜 Update History ")

    def _blink_loop(self):
        if not self.is_blinking:
            return
        self.blink_state = not self.blink_state
        blink_text = " 🔔 [NEW ALERTS] Update History " if self.blink_state else " 📜 Update History "
        self.notebook.tab(self.tab_history, text=blink_text)
        self.after(500, self._blink_loop)

    def set_inspector_text(self, text_content):
        self.insp_text.config(state="normal")
        self.insp_text.delete("1.0", tk.END)
        self.insp_text.insert("1.0", text_content)
        self.insp_text.config(state="disabled")

    def show_copy_menu(self, event, target_tree):
        row_id = target_tree.identify_row(event.y)
        if row_id:
            target_tree.selection_set(row_id)
            vals = target_tree.item(row_id)['values']
            row_str = " | ".join(str(v).replace('\n', ' ') for v in vals)

            menu = tk.Menu(self, tearoff=0)
            menu.add_command(label="📋 Copy Selected Row Details", command=lambda: self.copy_to_clipboard(row_str))
            menu.tk_popup(event.x_root, event.y_root)

    def copy_to_clipboard(self, text_to_copy):
        self.clipboard_clear()
        self.clipboard_append(text_to_copy)
        self.update()
        self.status_label.config(text="📋 Copied row to clipboard!", fg="#4ade80")

    def setup_resolved_table(self, parent_frame):
        self.resolved_tree = ttk.Treeview(
            parent_frame,
            columns=("ADEN_Dept", "Dealer", "Contractor", "BG_No_Date", "Amount", "Expiry_Date", "Remarks"),
            show="headings"
        )

        self.resolved_tree.heading("ADEN_Dept", text="ADEN / Dept.")
        self.resolved_tree.heading("Dealer", text="Dealer Name")
        self.resolved_tree.heading("Contractor", text="Contractor / Firm Name")
        self.resolved_tree.heading("BG_No_Date", text="Bank Guarantee No.")
        self.resolved_tree.heading("Amount", text="BG Amount")
        self.resolved_tree.heading("Expiry_Date", text="Validity Date")
        self.resolved_tree.heading("Remarks", text="Staff Valid Remarks")

        self.resolved_tree.column("ADEN_Dept", width=120, minwidth=100, anchor="center", stretch=True)
        self.resolved_tree.column("Dealer", width=130, minwidth=110, anchor="center", stretch=True)
        self.resolved_tree.column("Contractor", width=220, minwidth=160, anchor="w", stretch=True)
        self.resolved_tree.column("BG_No_Date", width=160, minwidth=130, anchor="w", stretch=True)
        self.resolved_tree.column("Amount", width=125, minwidth=100, anchor="e", stretch=True)
        self.resolved_tree.column("Expiry_Date", width=110, minwidth=90, anchor="center", stretch=True)
        self.resolved_tree.column("Remarks", width=320, minwidth=220, anchor="w", stretch=True)

        res_scroll = ttk.Scrollbar(parent_frame, orient="vertical", command=self.resolved_tree.yview)
        self.resolved_tree.configure(yscrollcommand=res_scroll.set)

        self.resolved_tree.pack(side="left", fill="both", expand=True)
        res_scroll.pack(side="right", fill="y")

        self.resolved_tree.tag_configure("RESOLVED", background="#dcfce7", foreground="#166534")
        self.resolved_tree.bind("<<TreeviewSelect>>", self.on_resolved_row_select)
        self.resolved_tree.bind("<Button-3>", lambda e: self.show_copy_menu(e, self.resolved_tree))

    def render_resolved_table(self, df_resolved):
        for item in self.resolved_tree.get_children():
            self.resolved_tree.delete(item)

        if df_resolved.empty:
            return

        for _, row in df_resolved.iterrows():
            aden_val = clean_cell_value(row.get(self.cols_map['aden'], 'N/A'))
            dealer_val = clean_cell_value(row.get(self.cols_map['dealer'], 'N/A'))
            contractor_raw = clean_cell_value(row.get(self.cols_map['contractor'], 'N/A'))
            bg_5th_raw = clean_cell_value(row.get(self.cols_map['bg_no'], 'N/A'))
            amt = clean_cell_value(row.get(self.cols_map['amt'], 'N/A'))
            validity = format_indian_date(row.get(self.cols_map['date'], 'N/A'))
            remarks_val = clean_cell_value(row.get(self.cols_map['remarks'], 'N/A'))

            contractor_wrapped = self.wrap_text_for_cell(contractor_raw, max_chars=22)
            bg_5th_wrapped = self.wrap_text_for_cell(bg_5th_raw, max_chars=18)
            remarks_wrapped = self.wrap_text_for_cell(remarks_val, max_chars=32)

            item_id = self.resolved_tree.insert("", "end", values=(
                aden_val, dealer_val, contractor_wrapped, bg_5th_wrapped, amt, validity, remarks_wrapped
            ), tags=("RESOLVED",))
            self.resolved_tree.item(item_id, text=f"{contractor_raw}||{bg_5th_raw}||{remarks_val}")

    def update_summary_cards(self, df_source):
        if df_source.empty:
            self.card_critical.config(text="0")
            self.card_red.config(text="0")
            self.card_orange.config(text="0")
            self.card_yellow.config(text="0")
            self.card_total.config(text="0")
            return

        critical_cnt = len(df_source[df_source['Radar_Warning'].str.contains('CRITICAL OVERDUE', case=False, na=False)])
        red_cnt = len(df_source[df_source['Radar_Warning'].str.contains('URGENT', case=False, na=False)])
        orange_cnt = len(df_source[df_source['Radar_Warning'].str.contains('HIGH PRIORITY', case=False, na=False)])
        yellow_cnt = len(df_source[df_source['Radar_Warning'].str.contains('DUE FOR RENEWAL', case=False, na=False)])
        total_cnt = len(df_source)

        self.card_critical.config(text=str(critical_cnt))
        self.card_red.config(text=str(red_cnt))
        self.card_orange.config(text=str(orange_cnt))
        self.card_yellow.config(text=str(yellow_cnt))
        self.card_total.config(text=str(total_cnt))

    def on_tab_changed(self, event):
        current_tab = self.notebook.index(self.notebook.select())
        if current_tab != 2:
            self.delete_btn.pack_forget()
        elif current_tab == 2:
            self.stop_tab_blinking()
            if self.selected_history_item:
                self.delete_btn.pack(fill="both", expand=True)

        self.filter_table()

    def setup_history_table(self, parent_frame):
        self.history_tree = ttk.Treeview(
            parent_frame,
            columns=("SNo", "ADEN_Dept", "Dealer", "Contractor", "BG_No", "Amount", "Old_Date", "New_Date", "Updated_By"),
            show="headings"
        )

        self.history_tree.heading("SNo", text="S.No.")
        self.history_tree.heading("ADEN_Dept", text="ADEN / Dept.")
        self.history_tree.heading("Dealer", text="Dealer Name")
        self.history_tree.heading("Contractor", text="Contractor / Firm Name")
        self.history_tree.heading("BG_No", text="Bank Guarantee No.")
        self.history_tree.heading("Amount", text="BG Amount")
        self.history_tree.heading("Old_Date", text="Old Date")
        self.history_tree.heading("New_Date", text="New Extended Date")
        self.history_tree.heading("Updated_By", text="Updated On (IST)")

        self.history_tree.column("SNo", width=55, anchor="center")
        self.history_tree.column("ADEN_Dept", width=100, anchor="center")
        self.history_tree.column("Dealer", width=120, anchor="center")
        self.history_tree.column("Contractor", width=190, anchor="w")
        self.history_tree.column("BG_No", width=140, anchor="w")
        self.history_tree.column("Amount", width=115, anchor="e")
        self.history_tree.column("Old_Date", width=100, anchor="center")
        self.history_tree.column("New_Date", width=115, anchor="center")
        self.history_tree.column("Updated_By", width=140, anchor="center")

        hist_scroll = ttk.Scrollbar(parent_frame, orient="vertical", command=self.history_tree.yview)
        self.history_tree.configure(yscrollcommand=hist_scroll.set)

        self.history_tree.pack(side="left", fill="both", expand=True)
        hist_scroll.pack(side="right", fill="y")

        self.history_tree.tag_configure("NEW_ENTRY", background="#e0f2fe", foreground="#0369a1")

        self.history_tree.bind("<<TreeviewSelect>>", self.on_history_row_select)
        self.history_tree.bind("<Button-3>", lambda e: self.show_copy_menu(e, self.history_tree))

    def render_history_table(self, df_hist):
        self.df_history_cached = df_hist
        for item in self.history_tree.get_children():
            self.history_tree.delete(item)

        if df_hist.empty:
            return

        row_count = 1
        for idx, row in df_hist.iterrows():
            vals = list(row)
            if len(vals) >= 9:
                aden_val       = clean_cell_value(vals[1])
                dealer_val     = clean_cell_value(vals[2])
                contractor_val = clean_cell_value(vals[3])
                bg_no_val      = clean_cell_value(vals[4])
                amt_val        = format_indian_currency(vals[5])
                
                old_date_val   = format_indian_date(vals[6])
                new_date_val   = format_indian_date(vals[7])
                updated_on_val = format_indian_datetime(vals[8])

                raw_updated_on = vals[8]
                signature = f"{bg_no_val}_{new_date_val}_{raw_updated_on}"

                tags = ("NEW_ENTRY",) if signature in self.new_logs_detected else ()

                item_id = self.history_tree.insert("", "end", values=(
                    f"🆕 {row_count}" if signature in self.new_logs_detected else row_count, 
                    aden_val, 
                    dealer_val, 
                    self.wrap_text_for_cell(contractor_val, max_chars=22), 
                    self.wrap_text_for_cell(bg_no_val, max_chars=18), 
                    amt_val, 
                    old_date_val, 
                    new_date_val, 
                    updated_on_val
                ), tags=tags)
                
                self.history_tree.item(item_id, text=f"{contractor_val}||{bg_no_val}||{signature}")
                row_count += 1

    def animate_sync_icon(self):
        sync_frames = ["🔄", "🔁", "🔂", "🔄"]
        idx = 0
        while self.is_syncing:
            frame = sync_frames[idx % len(sync_frames)]
            self.refresh_btn.config(text=f"{frame} Syncing Ledger...", bg="#047857")
            idx += 1
            time.sleep(0.18)
        
        self.refresh_btn.config(text="🔄 Live Sync", bg="#059669", state="normal")

    def create_card(self, parent, title, value, fg_color, bg_color, col_idx, filter_key):
        card = tk.Frame(parent, bg=bg_color, highlightbackground=self.BORDER_LINE, highlightthickness=1, height=85, cursor="hand2")
        card.grid(row=0, column=col_idx, padx=4, pady=5, sticky="nsew")
        parent.grid_columnconfigure(col_idx, weight=1)

        lbl_title = tk.Label(card, text=title, font=("Segoe UI", 8, "bold"), fg=fg_color, bg=bg_color, cursor="hand2")
        lbl_title.pack(anchor="w", padx=10, pady=(8, 0))

        lbl_val = tk.Label(card, text=value, font=("Segoe UI", 20, "bold"), fg=self.TEXT_DARK, bg=bg_color, cursor="hand2")
        lbl_val.pack(anchor="w", padx=10)

        card.bind("<Button-1>", lambda event, k=filter_key: self.filter_by_card(k))
        lbl_title.bind("<Button-1>", lambda event, k=filter_key: self.filter_by_card(k))
        lbl_val.bind("<Button-1>", lambda event, k=filter_key: self.filter_by_card(k))

        return lbl_val

    def filter_by_card(self, filter_key):
        if self.raw_warnings_df.empty:
            return

        self.notebook.select(self.tab_active)
        self.search_var.set("")

        if filter_key == "ALL":
            self.current_filtered_active_df = self.raw_warnings_df.copy()
            self.render_tree_data(self.current_filtered_active_df)
            self.status_label.config(text="📊 Displaying: ALL PENDING RADARS", fg="#38bdf8")
        else:
            self.current_filtered_active_df = self.raw_warnings_df[
                self.raw_warnings_df['Radar_Warning'].str.contains(filter_key, case=False, na=False)
            ].copy()
            self.render_tree_data(self.current_filtered_active_df)
            self.status_label.config(text=f"📊 Filtered: {filter_key} ({len(self.current_filtered_active_df)} records)", fg="#facc15")

        self.update_summary_cards(self.current_filtered_active_df)

    def wrap_text_for_cell(self, text, max_chars=22):
        cleaned = clean_cell_value(text)
        if len(cleaned) <= max_chars:
            return cleaned
        
        words = cleaned.split(' ')
        lines = []
        current_line = ""

        for word in words:
            if len(current_line) + len(word) + 1 <= max_chars:
                current_line += (" " if current_line else "") + word
            else:
                lines.append(current_line)
                current_line = word
        if current_line:
            lines.append(current_line)
            
        return "\n".join(lines[:2])

    def autofit_and_wrap(self, event):
        total_width = event.width - 20
        if total_width < 800:
            return

        weights = {
            "ADEN_Dept": 0.08,
            "Dealer": 0.08,
            "Contractor": 0.18,
            "BG_No_Date": 0.15,
            "Amount": 0.12,
            "Expiry_Date": 0.09,
            "Days_Left": 0.06,
            "Radar": 0.24
        }

        for col, weight in weights.items():
            calc_width = int(total_width * weight)
            self.tree.column(col, width=max(calc_width, 50))

    def trigger_async_load(self):
        if self.is_syncing:
            return
        
        self.is_syncing = True
        self.refresh_btn.config(state="disabled")
        self.status_label.config(text="⏳ Syncing Live Executive Stream...", fg="#facc15")
        
        threading.Thread(target=self.animate_sync_icon, daemon=True).start()
        threading.Thread(target=self._async_fetch_worker, daemon=True).start()

    def _async_fetch_worker(self):
        try:
            df_active, df_resolved, master_bg_lookup, date_col, aden_col, dealer_col, contractor_col, bg_no_col, amt_col, remarks_col = fetch_sheet_stream_fast()
            df_hist = fetch_history_stream(master_bg_lookup)
            self.after(0, self._update_ui_after_fetch, df_active, df_resolved, date_col, aden_col, dealer_col, contractor_col, bg_no_col, amt_col, remarks_col, df_hist)
        except Exception as e:
            self.after(0, self._handle_fetch_error, str(e))

    def _update_ui_after_fetch(self, df_active, df_resolved, date_col, aden_col, dealer_col, contractor_col, bg_no_col, amt_col, remarks_col, df_hist):
        self.is_syncing = False
        
        self.cols_map = {
            'date': date_col, 'aden': aden_col, 'dealer': dealer_col,
            'contractor': contractor_col, 'bg_no': bg_no_col, 'amt': amt_col,
            'remarks': remarks_col
        }

        self.raw_warnings_df = df_active[df_active['Days_Remaining'] <= 90].sort_values('Days_Remaining')
        self.df_resolved_cached = df_resolved.copy()
        self.df_history_cached = df_hist.copy()

        self.new_logs_detected.clear()
        if not df_hist.empty:
            for _, row in df_hist.iterrows():
                vals = [clean_cell_value(v) for v in row]
                if len(vals) >= 9:
                    bg_no = vals[4]
                    new_ext = vals[7]
                    updated_on = vals[8]
                    sig = f"{bg_no}_{new_ext}_{updated_on}"
                    
                    if sig not in self.known_entries:
                        self.new_logs_detected.add(sig)
                        self.known_entries.add(sig)

            save_json_set(KNOWN_ENTRIES_FILE, self.known_entries)

        self.filter_table()

        # Update Stat Cards
        self.update_summary_cards(self.raw_warnings_df)

        if len(self.new_logs_detected) > 0:
            self.start_tab_blinking()
            self.status_label.config(text=f"🔔 {len(self.new_logs_detected)} New Extension Logs Detected! [{datetime.now().strftime('%H:%M:%S')}]", fg="#facc15")
        else:
            self.status_label.config(text=f"✅ Data Synced [{datetime.now().strftime('%H:%M:%S')}]", fg="#4ade80")

    def _handle_fetch_error(self, err_msg):
        self.is_syncing = False
        messagebox.showerror("Sync Error", f"Failed to sync sheet data:\n{err_msg}")
        self.status_label.config(text="❌ Sync failed.", fg="#f87171")

    def render_tree_data(self, df_to_render):
        for item in self.tree.get_children():
            self.tree.delete(item)

        for _, row in df_to_render.iterrows():
            bg_5th_raw = clean_cell_value(row.get(self.cols_map['bg_no'], 'N/A')) if self.cols_map['bg_no'] else 'N/A'
            
            warning = str(row['Radar_Warning'])
            if 'CRITICAL OVERDUE' in warning:
                tag = "CRITICAL_OVERDUE"
            elif 'URGENT' in warning:
                tag = "URGENT"
            elif 'HIGH PRIORITY' in warning:
                tag = "HIGH_PRIORITY"
            elif 'DUE FOR RENEWAL' in warning:
                tag = "RENEWAL_DUE"
            else:
                tag = ""

            contractor_raw = clean_cell_value(row.get(self.cols_map['contractor'], 'N/A')) if self.cols_map['contractor'] else 'N/A'
            amt = clean_cell_value(row.get(self.cols_map['amt'], 'N/A')) if self.cols_map['amt'] else 'N/A'
            validity = format_indian_date(row.get(self.cols_map['date'], 'N/A'))
            days = row['Days_Remaining']

            aden_val = clean_cell_value(row.get(self.cols_map['aden'], 'N/A')) if self.cols_map['aden'] else 'N/A'
            dealer_val = clean_cell_value(row.get(self.cols_map['dealer'], 'N/A')) if self.cols_map['dealer'] else 'N/A'
            remarks_val = clean_cell_value(row.get(self.cols_map['remarks'], 'N/A')) if self.cols_map['remarks'] else 'N/A'

            contractor_wrapped = self.wrap_text_for_cell(contractor_raw, max_chars=22)
            bg_5th_wrapped = self.wrap_text_for_cell(bg_5th_raw, max_chars=18)

            item_id = self.tree.insert("", "end", values=(aden_val, dealer_val, contractor_wrapped, bg_5th_wrapped, amt, validity, days, row['Radar_Warning']), tags=(tag,))
            self.tree.item(item_id, text=f"{contractor_raw}||{bg_5th_raw}||{remarks_val}")

    def filter_table(self, *args):
        """Universal Live Search Engine: Always filters against pristine master cached dataframes"""
        query = self.search_var.get().lower().strip()
        current_tab_idx = self.notebook.index(self.notebook.select())

        if current_tab_idx == 0:  # Pending Active BG Radar Tab
            if not query or self.raw_warnings_df.empty:
                self.current_filtered_active_df = self.raw_warnings_df.copy()
            else:
                self.current_filtered_active_df = self.raw_warnings_df[
                    self.raw_warnings_df.apply(lambda row: query in row.astype(str).str.lower().to_string(), axis=1)
                ].copy()
            
            self.render_tree_data(self.current_filtered_active_df)
            self.update_summary_cards(self.current_filtered_active_df)
            self.status_label.config(
                text="📊 Displaying: ALL PENDING RADARS" if not query else f"🔍 Search Result: ({len(self.current_filtered_active_df)} Pending BGs)",
                fg="#38bdf8" if not query else "#facc15"
            )

        elif current_tab_idx == 1:  # Validated BGs Tab
            if not query or self.df_resolved_cached.empty:
                self.current_filtered_resolved_df = self.df_resolved_cached.copy()
            else:
                self.current_filtered_resolved_df = self.df_resolved_cached[
                    self.df_resolved_cached.apply(lambda row: query in row.astype(str).str.lower().to_string(), axis=1)
                ].copy()

            self.render_resolved_table(self.current_filtered_resolved_df)
            self.status_label.config(
                text=f"📊 Displaying: ALL Validated BGs ({len(self.df_resolved_cached)})" if not query else f"🔍 Search Result: ({len(self.current_filtered_resolved_df)} Validated BGs)",
                fg="#38bdf8" if not query else "#facc15"
            )

        elif current_tab_idx == 2:  # Audit History Tab
            if not query or self.df_history_cached.empty:
                self.current_filtered_history_df = self.df_history_cached.copy()
            else:
                self.current_filtered_history_df = self.df_history_cached[
                    self.df_history_cached.apply(lambda row: query in row.astype(str).str.lower().to_string(), axis=1)
                ].copy()

            self.render_history_table(self.current_filtered_history_df)
            self.status_label.config(
                text=f"📊 Displaying: ALL Audit History ({len(self.df_history_cached)})" if not query else f"🔍 Search Result: ({len(self.current_filtered_history_df)} Audit Logs)",
                fg="#38bdf8" if not query else "#facc15"
            )

    def export_excel(self):
        """Universal Smart Excel Export"""
        current_tab_idx = self.notebook.index(self.notebook.select())
        query = self.search_var.get().strip()

        target_tree = self.tree if current_tab_idx == 0 else (self.resolved_tree if current_tab_idx == 1 else self.history_tree)
        
        if target_tree.get_children() == ():
            messagebox.showwarning("Export Warning", "No records currently displayed on active tab to export.")
            return

        tab_name_clean = "pending" if current_tab_idx == 0 else ("validated" if current_tab_idx == 1 else "audit_history")
        if query:
            safe_query = re.sub(r'[\\/*?:"<>|]', "", query).replace(' ', '_').lower()
            default_filename = f"bg_{safe_query}_{tab_name_clean}_report"
        else:
            default_filename = f"bg_{tab_name_clean}_report"

        file_path = filedialog.asksaveasfilename(
            initialfile=default_filename,
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")],
            title="Export Active Tab Executive Report"
        )
        if file_path:
            try:
                export_records = []
                for item in target_tree.get_children():
                    row_vals = list(target_tree.item(item)['values'])
                    
                    if current_tab_idx == 0:
                        row_vals[2] = str(row_vals[2]).replace('\n', ' ')
                        row_vals[3] = str(row_vals[3]).replace('\n', ' ')
                        raw_meta = target_tree.item(item)['text']
                        remarks_val = "N/A"
                        if "||" in raw_meta:
                            parts = raw_meta.split("||")
                            if len(parts) >= 3:
                                remarks_val = parts[2]
                        row_vals.append(remarks_val)
                    elif current_tab_idx in [1, 2]:
                        if len(row_vals) > 3:
                            row_vals[2] = str(row_vals[2]).replace('\n', ' ')
                            row_vals[3] = str(row_vals[3]).replace('\n', ' ')
                    
                    export_records.append(row_vals)

                if current_tab_idx == 0:
                    gui_columns = ["ADEN / Dept.", "Dealer Name", "Name of Contractor", "Bank Guarantee No.", "BG Amount", "Validity Date", "Days Left", "Executive Radar Status", "Staff Remarks"]
                    report_title = "INDIAN RAILWAYS — PENDING BG RADAR EXECUTIVE REPORT"
                elif current_tab_idx == 1:
                    gui_columns = ["ADEN / Dept.", "Dealer Name", "Name of Contractor", "Bank Guarantee No.", "BG Amount", "Validity Date", "Staff Remarks"]
                    report_title = "INDIAN RAILWAYS — VALIDATED & RESOLVED BGs REPORT"
                else:
                    gui_columns = ["S.No.", "ADEN / Dept.", "Dealer Name", "Contractor / Firm Name", "BG No.", "BG Amount", "Old Date", "New Extended Date", "Updated On (IST)"]
                    report_title = "INDIAN RAILWAYS — BANK GUARANTEE AUDIT LOG HISTORY"

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Executive_Report"

                col_count = len(gui_columns)
                end_col_letter = get_column_letter(col_count)

                ws.merge_cells(f"A1:{end_col_letter}1")
                title_cell = ws["A1"]
                title_cell.value = report_title
                title_cell.font = Font(name="Segoe UI", size=13, bold=True, color="FFFFFF")
                title_cell.fill = PatternFill(start_color="061D33", end_color="061D33", fill_type="solid")
                title_cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[1].height = 34

                ws.merge_cells(f"A2:{end_col_letter}2")
                sub_cell = ws["A2"]
                sub_cell.value = f"Report Generated (IST): {datetime.now().strftime('%d/%m/%Y %I:%M %p')} | Total Records: {len(export_records)}"
                sub_cell.font = Font(name="Segoe UI", size=9, italic=True, color="64748B")
                sub_cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[2].height = 20

                ws.append([])
                ws.row_dimensions[3].height = 10
                
                ws.append(gui_columns)
                ws.row_dimensions[4].height = 28

                header_fill = PatternFill(start_color="0F2942", end_color="0F2942", fill_type="solid")
                header_font = Font(name="Segoe UI", size=11, bold=True, color="D97706")
                header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

                thin_border = Border(
                    left=Side(style='thin', color='CBD5E1'),
                    right=Side(style='thin', color='CBD5E1'),
                    top=Side(style='thin', color='CBD5E1'),
                    bottom=Side(style='thin', color='CBD5E1')
                )

                for col_num in range(1, col_count + 1):
                    cell = ws.cell(row=4, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_align
                    cell.border = thin_border

                fill_critical = PatternFill(start_color="FFE4E6", end_color="FFE4E6", fill_type="solid")
                font_critical = Font(name="Segoe UI", size=10, bold=True, color="991B1B")

                fill_urgent = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                font_urgent = Font(name="Segoe UI", size=10, bold=True, color="991B1B")

                fill_high = PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid")
                font_high = Font(name="Segoe UI", size=10, bold=True, color="9A3412")

                fill_due = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
                font_due = Font(name="Segoe UI", size=10, bold=True, color="854D0E")

                fill_valid_remarks = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
                font_valid_remarks = Font(name="Segoe UI", size=10, bold=True, color="166534")

                cell_font = Font(name="Segoe UI", size=10)
                wrap_align = Alignment(vertical="center", wrap_text=True)
                center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
                right_align = Alignment(horizontal="right", vertical="center", wrap_text=True)

                for idx, row in enumerate(export_records, start=5):
                    ws.append(list(row))
                    ws.row_dimensions[idx].height = 26

                    for col_idx in range(1, col_count + 1):
                        cell = ws.cell(row=idx, column=col_idx)
                        cell.font = cell_font
                        cell.border = thin_border

                        if current_tab_idx == 0:
                            if col_idx in [1, 2, 6, 7]:
                                cell.alignment = center_align
                            elif col_idx == 5:
                                cell.alignment = right_align
                            else:
                                cell.alignment = wrap_align

                            status_str = str(row[7])
                            if col_idx in [8, 9]:
                                if 'CRITICAL OVERDUE' in status_str:
                                    cell.fill = fill_critical
                                    cell.font = font_critical
                                elif 'URGENT' in status_str:
                                    cell.fill = fill_urgent
                                    cell.font = font_urgent
                                elif 'HIGH PRIORITY' in status_str:
                                    cell.fill = fill_high
                                    cell.font = font_high
                                elif 'DUE FOR RENEWAL' in status_str:
                                    cell.fill = fill_due
                                    cell.font = font_due

                        elif current_tab_idx == 1:
                            if col_idx in [1, 2, 6]:
                                cell.alignment = center_align
                            elif col_idx == 5:
                                cell.alignment = right_align
                            else:
                                cell.alignment = wrap_align

                            if col_idx == 7:
                                cell.fill = fill_valid_remarks
                                cell.font = font_valid_remarks

                        else:
                            if col_idx in [1, 2, 3, 7, 8, 9]:
                                cell.alignment = center_align
                            elif col_idx == 6:
                                cell.alignment = right_align
                            else:
                                cell.alignment = wrap_align

                for col in ws.columns:
                    max_len = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        if cell.row < 4:
                            continue
                        val_str = str(cell.value or '')
                        if len(val_str) > max_len:
                            max_len = len(val_str)
                    calc_width = min(max(max_len + 4, 15), 45)
                    ws.column_dimensions[col_letter].width = calc_width

                ws.auto_filter.ref = f"A4:{end_col_letter}{len(export_records) + 4}"
                wb.save(file_path)

                messagebox.showinfo(
                    "Export Successful", 
                    f"Executive report saved successfully at:\n{file_path}"
                )

            except Exception as e:
                messagebox.showerror("Export Error", f"Failed to save Excel file:\n{e}")

    def on_row_select(self, event):
        """Inspector View for Active BG Radar Tab"""
        self.selected_history_item = None
        self.delete_btn.pack_forget()
        selected_items = self.tree.selection()
        if not selected_items:
            return

        selected_id = selected_items[0]
        vals = self.tree.item(selected_id)['values']
        raw_meta = self.tree.item(selected_id)['text']

        contractor_full, bg_full, remarks_full = vals[2], vals[3], "N/A"
        if "||" in raw_meta:
            parts = raw_meta.split("||")
            contractor_full = parts[0]
            bg_full = parts[1] if len(parts) > 1 else vals[3]
            remarks_full = parts[2] if len(parts) > 2 else "N/A"

        detail_text = (
            f"📋 PENDING BG RECORD INSPECTOR\n"
            f"───────────────────\n"
            f"👤 Active Staff: {self.active_user}\n\n"
            f"📌 Dept.:  {vals[0]}\n\n"
            f"👤 Dealer:  {vals[1]}\n\n"
            f"🏢 Contractor:  {contractor_full}\n\n"
            f"📄 Bank Guarantee No.:  {bg_full}\n\n"
            f"💰 BG Amount:  {vals[4]}\n\n"
            f"📅 Validity Expiry Date:  {vals[5]}\n\n"
            f"⏳ Days Remaining:  {vals[6]} Days\n\n"
            f"🚨 Executive Radar Status:\n  {vals[7]}\n\n"
            f"📝 Staff Remarks:\n  {remarks_full}"
        )
        self.set_inspector_text(detail_text)

    def on_resolved_row_select(self, event):
        """Inspector View for Validated/Resolved BGs Tab"""
        self.selected_history_item = None
        self.delete_btn.pack_forget()
        selected_items = self.resolved_tree.selection()
        if not selected_items:
            return

        selected_id = selected_items[0]
        vals = self.resolved_tree.item(selected_id)['values']
        raw_meta = self.resolved_tree.item(selected_id)['text']

        contractor_full, bg_full, remarks_full = vals[2], vals[3], vals[6]
        if "||" in raw_meta:
            parts = raw_meta.split("||")
            contractor_full = parts[0]
            bg_full = parts[1] if len(parts) > 1 else vals[3]
            remarks_full = parts[2] if len(parts) > 2 else vals[6]

        detail_text = (
            f"✅ VALIDATED BG INSPECTOR\n"
            f"───────────────────\n"
            f"👤 Active Staff: {self.active_user}\n\n"
            f"📌 Dept.:  {vals[0]}\n\n"
            f"👤 Dealer:  {vals[1]}\n\n"
            f"🏢 Contractor:  {contractor_full}\n\n"
            f"📄 Bank Guarantee No.:  {bg_full}\n\n"
            f"💰 BG Amount:  {vals[4]}\n\n"
            f"📅 Validity Expiry Date:  {vals[5]}\n\n"
            f"📝 Staff Valid Remarks:\n  {remarks_full}"
        )
        self.set_inspector_text(detail_text)

    def on_history_row_select(self, event):
        """Inspector View for Update History Tab"""
        selected_items = self.history_tree.selection()
        if not selected_items:
            self.selected_history_item = None
            self.delete_btn.pack_forget()
            return

        self.selected_history_item = selected_items[0]
        vals = self.history_tree.item(self.selected_history_item)['values']
        raw_meta = self.history_tree.item(self.selected_history_item)['text']

        if "||" in raw_meta:
            parts = raw_meta.split("||")
            contractor_full, bg_full = parts[0], parts[1]
        else:
            contractor_full, bg_full = vals[3], vals[4]

        detail_text = (
            f"📜 AUDIT LOG INSPECTOR\n"
            f"───────────────────\n"
            f"👤 Active Staff: {self.active_user}\n\n"
            f"📌 Dept.:  {vals[1]}\n\n"
            f"👤 Dealer:  {vals[2]}\n\n"
            f"🏢 Contractor/Firm:  {contractor_full}\n\n"
            f"📄 Bank Guarantee No.:  {vals[4]}\n\n"
            f"💰 BG Amount:  {vals[5]}\n\n"
            f"⏳ Previous Old Date:  {vals[6]}\n\n"
            f"📅 New Extended Date:  {vals[7]}\n\n"
            f"🕒 Updated On (IST):\n  {vals[8]}"
        )
        self.set_inspector_text(detail_text)
        self.delete_btn.pack(fill="both", expand=True)

    def delete_history_entry(self):
        """Deletes selected entry permanently from UI and persists deletion across syncs"""
        if not self.selected_history_item:
            messagebox.showwarning("Selection Warning", "Please select a history entry row to delete.")
            return

        vals = self.history_tree.item(self.selected_history_item)['values']
        raw_meta = self.history_tree.item(self.selected_history_item)['text']
        s_no = str(vals[0]).replace('🆕 ', '')
        bg_no = vals[4]

        confirm = messagebox.askyesno(
            "Confirm Deletion",
            f"Are you sure you want to permanently remove History Log S.No. {s_no}?\n(BG No: {bg_no})"
        )

        if confirm:
            if "||" in raw_meta:
                parts = raw_meta.split("||")
                if len(parts) >= 3:
                    signature = parts[2]
                    self.deleted_blacklist.add(signature)
                    save_json_set(DELETED_LOGS_FILE, self.deleted_blacklist)

            self.history_tree.delete(self.selected_history_item)
            self.selected_history_item = None
            self.delete_btn.pack_forget()
            self.set_inspector_text("Click on any record row to inspect full details.")

            remaining_children = self.history_tree.get_children()
            for idx, child_id in enumerate(remaining_children, start=1):
                child_vals = list(self.history_tree.item(child_id)['values'])
                s_no_str = str(child_vals[0])
                prefix = "🆕 " if "🆕" in s_no_str else ""
                child_vals[0] = f"{prefix}{idx}"
                self.history_tree.item(child_id, values=child_vals)

            messagebox.showinfo("Success", f"History Entry S.No. {s_no} deleted permanently.")


if __name__ == "__main__":
    app = ExecutiveRailwaysPortal()
    app.mainloop()